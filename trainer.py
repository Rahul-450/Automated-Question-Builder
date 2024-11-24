# trainer.py

import csv
import os
from issues import add_issue
from QA_Generator_modular import generate_questions
from openai import AzureOpenAI
import yaml
import json
import pandas as pd
from QA_generator_KB import generate_questions_from_kb


# Reading Config File
current_path = os.getcwd()
full_config_path = "config.yaml"
config_values = yaml.safe_load(open(full_config_path))

AZURE_OPENAI_ENDPOINT = config_values['AZURE_OPENAI_ENDPOINT']
AZURE_OPENAI_KEY = config_values['AZURE_OPENAI_API_KEY']
OPENAI_API_VERSION = config_values['OPENAI_API_VERSION']
DEPLOYMENT_NAME=config_values['DEPLOYMENT_NAME']
QUESTION_TYPES = ["MCQ", "Case Study", "True or False", "Fill Ups", "Find the Code Output", "Find the Error"]
question_bank_directory=r"C:\Users\2000108450\OneDrive - Hexaware Technologies\Desktop\AutomatedQuestionBuilder\data\question_banks"
all_question_bank_details_path=r"C:\Users\2000108450\OneDrive - Hexaware Technologies\Desktop\AutomatedQuestionBuilder\data\question_banks_details.json"
kbs_directory=r"C:\Users\2000108450\OneDrive - Hexaware Technologies\Desktop\AutomatedQuestionBuilder\data\kbs"
try:
    client = AzureOpenAI(
        azure_endpoint=AZURE_OPENAI_ENDPOINT,
        api_key=AZURE_OPENAI_KEY,
        api_version=OPENAI_API_VERSION
    )
except:
    print("Azure OpenAI connection error.")

CURRICULUM_FILE = 'data/curriculum.csv'
QUESTION_BANK_DIR = 'data/question_banks/'


# Load Question Bank from file
def load_question_bank(file_path):
    if os.path.exists(file_path):
        with open(file_path, 'r') as file:
            return json.load(file)
    else:
        return {"questions": []}

question_bank_path = r"C:\Users\2000108450\OneDrive - Hexaware Technologies\Desktop\AutomatedQuestionBuilder\data\question_bank.json"
question_bank = load_question_bank(question_bank_path)

def upload_curriculum():
    print("\nUpload Curriculum")
    filepath = input("Enter path to curriculum CSV file: ")
    if not os.path.isfile(filepath):
        print("File does not exist.")
        return
    try:
        with open(filepath, 'r') as f_in, open(CURRICULUM_FILE, 'w') as f_out:
            content = f_in.read()
            f_out.write(content)
        print("Curriculum uploaded successfully.")
    except Exception as e:
        print(f"An error occurred: {e}")



def extract_subtopics(file_path,Topic_name):
    subtopics = []
    try:
        
        # reader = pd.read_csv(file_path)
        reader = pd.read_excel(file_path)
        # Assuming subtopics are in the first column after the header
        sys_prompt="""
### Role:
AI Assistant

### Instructions:
1. You will be given the content of a csv file and a Topic name.
2. Go through the whole content and extract the subtopics from the given CSV file content which are related to the given Topic.
3. subtopics should be as a list of subtopics. 
4. Strictly return the output as a JSON format where key should be subtopics and value should be list of subtopics.

### Example Topic Name:
Full STACK DEVELOPER

### Example Output Response:
 {"subtopics": ["c", "Python", "c#", "HTML", "CSS",] }

"""
        user_prompt=f"""
### Role:
AI Assistant

### Instructions:
1. You will be given the content of a csv file and a Topic name.
2. Go through the whole content and extract the subtopics from the given topic which are related to the given Topic
3. subtopics should be as a list of subtopics. 
4. Strictly return the output as a JSON format where key should be subtopics and value should be list of subtopics.

###Topic Name:
{Topic_name}

### CSV File Content:
{reader}

"""
        print(f"user_prompt is {user_prompt}")
        messages = [
            {"role": "system", "content": sys_prompt},
            {"role": "user", "content": user_prompt}
        ]
        response = client.chat.completions.create(
                model=DEPLOYMENT_NAME,
                messages=messages,
                response_format={"type": "json_object"},
                temperature=0.7,
                max_tokens=4096
            )
        result = response.choices[0].message.content
        print(f"result is {result}")
    
        # Ensure the result is valid JSON
        try:
            # logger.info("API Response:", result)  # Debug: logger.info the API response
            json_output= json.loads(result)
            subtopics_list=json_output["subtopics"]
            return subtopics_list
        except json.JSONDecodeError as e:
            print("JSON Decode Error:", str(e))
            # logger.info("Result received:", result)
            return  None
        
    except Exception as e:
        print(f"An error occurred while extracting subtopics: {e}")
    return None



def extract_subtopics_for_kb(content,Topic_name):
    subtopics = []
    try:

        sys_prompt="""
### Role:
AI Assistant

### Instructions:
1. You will be given the content of a file and a Topic name.
2. Go through the whole content and extract the subtopics from the given file content which are related to the given Topic.
3. subtopics should be as a list of subtopics. 
4. Strictly return the output as a JSON format where key should be subtopics and value should be list of subtopics.

### Example Topic Name:
Full STACK DEVELOPER

### Example Output Response:
 {"subtopics": ["c", "Python", "c#", "HTML", "CSS",] }

"""
        user_prompt=f"""
### Role:
AI Assistant

### Instructions:
1. You will be given the content of a file and a Topic name.
2. Go through the whole content and extract the subtopics from the given file content which are related to the given Topic.
3. subtopics should be as a list of subtopics. 
4. Strictly return the output as a JSON format where key should be subtopics and value should be list of subtopics.

###Topic Name:
{Topic_name}

### File Content:
{content}

"""
        print(f"user_prompt is {user_prompt}")
        messages = [
            {"role": "system", "content": sys_prompt},
            {"role": "user", "content": user_prompt}
        ]
        response = client.chat.completions.create(
                model=DEPLOYMENT_NAME,
                messages=messages,
                response_format={"type": "json_object"},
                temperature=0.7,
                max_tokens=4096
            )
        result = response.choices[0].message.content
        print(f"result is {result}")
    
        # Ensure the result is valid JSON
        try:
            # logger.info("API Response:", result)  # Debug: logger.info the API response
            json_output= json.loads(result)
            subtopics_list=json_output["subtopics"]
            return subtopics_list
        except json.JSONDecodeError as e:
            print("JSON Decode Error:", str(e))
            # logger.info("Result received:", result)
            return  None
        
    except Exception as e:
        print(f"An error occurred while extracting subtopics: {e}")
    return None


def smart_question_generator(choice):
    print("\n Question Generator")
    topic = input("Enter Topic Name: ")
    print("Please upload the curriculum file in CSV format.")
    filepath = input("Enter path to curriculum CSV file: ")
    if not os.path.isfile(filepath):
        print("File does not exist.")
        return
    subtopics = extract_subtopics(filepath,topic)
    if not subtopics:
        print("No subtopics found in the curriculum.")
        return
    print("\nAvailable Subtopics:")
    for idx, subtopic in enumerate(subtopics, start=1):
        print(f"{idx}. {subtopic}")
    selected_indices = input("Enter the numbers of the subtopics to include (comma-separated): ")
    try:
        indices = [int(idx.strip()) - 1 for idx in selected_indices.split(',')]
        selected_subtopics = [subtopics[i] for i in indices if 0 <= i < len(subtopics)]
    except ValueError:
        print("Invalid selection.")
        return
    if not selected_subtopics:
        print("No valid subtopics selected.")
        return
    try:
        total_easy = int(input("Enter number of easy questions: "))
        total_medium = int(input("Enter number of medium questions: "))
        total_hard = int(input("Enter number of hard questions: "))
        highest_questions=max(total_easy,total_hard,total_medium)
        if highest_questions == total_easy:
            Difficulity="begginer"
        elif highest_questions == total_medium:
            Difficulity="Intermediate"
        elif highest_questions == total_hard:
            Difficulity = "Advanced"


    except ValueError:
        print("Please enter valid numbers for difficulty levels.")
        return
    total_questions = total_easy + total_medium + total_hard
    print("\nAvailable Question Types:")
    for idx, qtype in enumerate(QUESTION_TYPES, start=1):
        print(f"{idx}. {qtype}")
    selected_qtypes_indices = input("Enter the numbers of the question types to include (comma-separated): ")
    try:
        qtype_indices = [int(idx.strip()) - 1 for idx in selected_qtypes_indices.split(',')]
        selected_qtypes = [QUESTION_TYPES[i] for i in qtype_indices if 0 <= i < len(QUESTION_TYPES)]
    except ValueError:
        print("Invalid selection.")
        return
    if not selected_qtypes:
        print("No valid question types selected.")
        return
    user_input = {
        "topic": topic,
        "subtopics": selected_subtopics,
        "total_number_of_questions": total_questions,
        "difficulty_levels": {"easy": total_easy, "medium": total_medium, "hard": total_hard},
        "types": selected_qtypes,
    }
    all_questions, unique_questions_set, final_difficulty_counts = generate_questions(user_input, question_bank, question_bank_path, choice)
    question_bank_path_to_save=save_data_to_json(all_questions,topic)
    save_qb_details(question_bank_path_to_save,topic,subtopics,Difficulity)
    print("Question bank has been generated successfully.")


def generate_from_kb():
    print("\nGenerate Question Bank from Knowledge Base")
    topic = input("Enter Topic Name: ")
    print("Please upload the knowledge base file (.txt, .pdf, .docx, .xlsx, .csv).")
    filepath = input("Enter path to knowledge base file: ")
    if not os.path.isfile(filepath):
        print("File does not exist.")
        return
    content = extract_content(filepath)
    if not content:
        print("Failed to extract content from the file.")
        return
    subtopics = extract_subtopics_for_kb(content,topic)
    if not subtopics:
        print("No subtopics could be extracted from the content.")
        return
    subtopics = list(set(subtopics))  # Remove duplicates
    print("\nExtracted Subtopics:")
    for idx, subtopic in enumerate(subtopics, start=1):
        print(f"{idx}. {subtopic}")
    selected_indices = input("Enter the numbers of the subtopics to include (comma-separated): ")
    try:
        indices = [int(idx.strip()) - 1 for idx in selected_indices.split(',')]
        selected_subtopics = [subtopics[i] for i in indices if 0 <= i < len(subtopics)]
    except ValueError:
        print("Invalid selection.")
        return
    if not selected_subtopics:
        print("No valid subtopics selected.")
        return
    try:
        total_easy = int(input("Enter number of easy questions: "))
        total_medium = int(input("Enter number of medium questions: "))
        total_hard = int(input("Enter number of hard questions: "))
    except ValueError:
        print("Please enter valid numbers for difficulty levels.")
        return
    total_questions = total_easy + total_medium + total_hard
    print("\nAvailable Question Types:")
    for idx, qtype in enumerate(QUESTION_TYPES, start=1):
        print(f"{idx}. {qtype}")
    selected_qtypes_indices = input("Enter the numbers of the question types to include (comma-separated): ")
    try:
        qtype_indices = [int(idx.strip()) - 1 for idx in selected_qtypes_indices.split(',')]
        selected_qtypes = [QUESTION_TYPES[i] for i in qtype_indices if 0 <= i < len(QUESTION_TYPES)]
    except ValueError:
        print("Invalid selection.")
        return
    if not selected_qtypes:
        print("No valid question types selected.")
        return
    user_input = {
        "topic": topic,
        "subtopics": selected_subtopics,
        "total_number_of_questions": total_questions,
        "difficulty_levels": {"easy": total_easy, "medium": total_medium, "hard": total_hard},
        "types": selected_qtypes,
    }
    # Note: Assuming `question_builder_kb` is a function you have implemented
    all_generated_questions, final_difficulty_counts = generate_questions_from_kb(user_input, content)
    question_bank_path_to_save=save_data_to_json(all_generated_questions,topic)
    save_qb_details(question_bank_path_to_save,topic,subtopics)

    print("Question bank has been generated successfully.")



def save_qb_details(question_bank_path_to_save,topic, subtopics,difficulity):
    # Read the existing JSON data from the file
    
    if os.path.exists(all_question_bank_details_path) and os.path.getsize(all_question_bank_details_path) > 0:
        with open(all_question_bank_details_path, 'r') as file:
            data = json.load(file)
    else:
        data=[]
    # Create a new object to add
    new_object = {
        "topic": topic,
        "file_path": question_bank_path_to_save,
        "subtopics": subtopics,
        "Difficulity":difficulity
    }
    
    # Add the new object to the existing data
    data.append(new_object)
    
    # Write the updated data back to the file
    with open(all_question_bank_details_path, 'w') as file:
        json.dump(data, file, indent=4)

def save_data_to_json(data,topic):
    """
    Save the given data to a JSON file at the specified file path.

    :param data: The data to be saved (list of dictionaries).
    """
    try:
        file_name=f"{topic}.json"
        question_bank_path_to_save=os.path.join(question_bank_directory,file_name)
        with open(question_bank_path_to_save, 'w') as json_file:
            json.dump(data, json_file, indent=4)
        print(f"Data successfully saved to {question_bank_path_to_save}")
        return question_bank_path_to_save
    except Exception as e:
        print(f"An error occurred while saving data: {e}")
        return None

from PyPDF2 import PdfReader
from docx import Document
import openpyxl

def extract_content(file_path):
    extension = os.path.splitext(file_path)[1].lower()
    content = ""
    try:
        if extension == '.txt':
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
        elif extension == '.pdf':
            reader = PdfReader(file_path)
            for page in reader.pages:
                content += page.extract_text()
        elif extension == '.docx':
            doc = Document(file_path)
            for para in doc.paragraphs:
                content += para.text + '\n'
        elif extension == '.xlsx':
            wb = openpyxl.load_workbook(file_path)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    content += ' '.join([str(cell) if cell is not None else '' for cell in row]) + '\n'
        elif extension == '.csv':
            with open(file_path, 'r', encoding='utf-8') as csvfile:
                reader = csv.reader(csvfile)
                for row in reader:
                    content += ' '.join(row) + '\n'
        else:
            print("Unsupported file format.")
        return content
    except Exception as e:
        print(f"An error occurred while reading the file: {e}")
        return None



def generate_question_bank():
    while True:
        print("\nGenerate Question Bank")
        print("1. Smart Question Generator")
        print("2. Advanced AI Generator")
        print("3. Generate From KB")
        print("4. Back to Trainer Menu")
        choice = input("Enter your choice: ")
        if choice == '1':
            smart_question_generator("smart_generation")
        elif choice == '2':
            smart_question_generator("AI")
        elif choice == '3':
            generate_from_kb()
        elif choice == '4':
            break
        else:
            print("Invalid choice. Please select from the menu options.")

def upload_question_bank():
    print("\nUpload Question Bank")
    file_path = input("Enter the file path of the question bank: ")
    if not os.path.isfile(file_path):
        print("File does not exist. Please try again.")
        return

    try:
        # Get the filename and construct the destination path
        filename = os.path.basename(file_path)
        destination = os.path.join(question_bank_directory, filename)

        # Check if a file with the same name already exists
        if os.path.exists(destination):
            choice = input("A file with the same name already exists. Do you want to overwrite it? (yes/no): ").lower()
            if choice != 'yes':
                print("Upload cancelled.")
                return

        # Copy the file to the destination directory
        with open(file_path, 'rb') as f_in, open(destination, 'wb') as f_out:
            f_out.write(f_in.read())
        print(f"Question bank '{filename}' has been uploaded successfully.")
    except Exception as e:
        print(f"An error occurred while uploading the question bank: {e}")

def upload_kb():
    print("\nUpload Knowledge Base (KB)")
    file_path = input("Enter the file path of the knowledge base: ")
    if not os.path.isfile(file_path):
        print("File does not exist. Please try again.")
        return

    try:
        # Get the filename and construct the destination path
        filename = os.path.basename(file_path)
        destination = os.path.join(kbs_directory, filename)

        # Check if a file with the same name already exists
        if os.path.exists(destination):
            choice = input("A file with the same name already exists. Do you want to overwrite it? (yes/no): ").lower()
            if choice != 'yes':
                print("Upload cancelled.")
                return

        # Copy the file to the destination directory
        with open(file_path, 'rb') as f_in, open(destination, 'wb') as f_out:
            f_out.write(f_in.read())
        print(f"Knowledge base '{filename}' has been uploaded successfully.")
    except Exception as e:
        print(f"An error occurred while uploading the knowledge base: {e}")

def trainer_menu(username):
    while True:
        print("\nTrainer Menu:")
        print("1. Upload Curriculum")
        print("2. Generate Question Bank")
        print("3. Upload Question Bank")
        print("4. Upload Knowledge Base")
        print("5. Raise Issue")
        print("6. Logout")
        choice = input("Enter your choice: ")

        if choice == '1':
            upload_curriculum()
        elif choice == '2':
            generate_question_bank()
        elif choice == '3':
            upload_question_bank()
        elif choice == '4':
            upload_kb()
        elif choice == '5':
            add_issue(username)
        elif choice == '6':
            print("Logging out...")
            break
        else:
            print("Invalid choice. Please select from the menu options.")